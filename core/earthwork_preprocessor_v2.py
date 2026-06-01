from __future__ import annotations

from collections import deque
from contextlib import contextmanager
from dataclasses import dataclass
from pathlib import Path
import re
import shutil
from typing import Any
import warnings

import geopandas as gpd
from openpyxl import load_workbook
from openpyxl.worksheet._writer import ALL_TEMP_FILES, WorksheetWriter
import pandas as pd

from core.temporary_occupancy import TemporaryOccupancySelection, calculate_temporary_width, get_default_selection
from utils.helpers import (
    coerce_to_float,
    ensure_directory,
    now_timestamp,
    override_temp_directory,
    prepare_runtime_temp_directory,
)
from utils.logger import get_log_messages


class EarthworkPrepError(ValueError):
    pass


LIST_SHEET_PATTERN = re.compile(r"^관로토공LIST\((?P<diameter_group>.+)\)$")
PROFILE_SHEET_PATTERN = re.compile(r"^토적산출표\((?P<line_name>.+)\)$")
STRUCTURE_LIST_PATTERN = re.compile(r"^변실LIST$")
SECTION_SHEET_PATTERN = re.compile(r"^(배수|가압장|감압변실|유량계|공기변|물탱크)")
PIPE_SHEET_PATTERN = re.compile(r"관수량")
BH_PATTERN = re.compile(r"BH\s*(?P<value>\d+(?:\.\d+)?)", re.IGNORECASE)
PARENS_PATTERN = re.compile(r"\((?P<value>[^()]+)\)")
DIAMETER_HINT_PATTERN = re.compile(r"(?P<mm>\d+(?:\.\d+)?)")
DIAMETER_PAIR_PATTERN = re.compile(r"\((?P<inner>\d+(?:\.\d+)?)\)")


@dataclass
class EarthworkPrepInputs:
    workbook_path: str
    output_dir: str
    pipe_workbook_path: str
    pipe_shapefile_path: str


@dataclass
class EarthworkPrepResult:
    segment_table: pd.DataFrame
    final_input_table: pd.DataFrame
    line_registry_table: pd.DataFrame
    section_reference_table: pd.DataFrame
    structure_table: pd.DataFrame
    pipe_feature_table: pd.DataFrame
    review_table: pd.DataFrame
    output_files: dict[str, str]


@contextmanager
def _suppress_openpyxl_cleanup_permission_error():
    original_cleanup = WorksheetWriter.cleanup

    def safe_cleanup(self) -> None:
        try:
            original_cleanup(self)
        except PermissionError:
            if getattr(self, "out", None) in ALL_TEMP_FILES:
                try:
                    ALL_TEMP_FILES.remove(self.out)
                except ValueError:
                    pass

    WorksheetWriter.cleanup = safe_cleanup
    try:
        yield
    finally:
        WorksheetWriter.cleanup = original_cleanup


def _load_prep_config(config: dict[str, Any]) -> dict[str, Any]:
    prep_config = config.get("earthwork_preprocessor")
    if not isinstance(prep_config, dict):
        raise EarthworkPrepError("earthwork_preprocessor 설정이 없습니다.")
    return prep_config


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _cell_value(row: tuple[Any, ...], one_based_index: int) -> Any:
    zero_based_index = one_based_index - 1
    if zero_based_index < 0 or zero_based_index >= len(row):
        return None
    return row[zero_based_index]


def _visible_worksheets(workbook):
    return [worksheet for worksheet in workbook.worksheets if worksheet.sheet_state == "visible"]


def _sheet_type(sheet_name: str) -> str | None:
    if LIST_SHEET_PATTERN.search(sheet_name):
        return "earthwork_list"
    if PROFILE_SHEET_PATTERN.search(sheet_name):
        return "profile"
    if STRUCTURE_LIST_PATTERN.search(sheet_name):
        return "structure_list"
    if PIPE_SHEET_PATTERN.search(sheet_name):
        return "pipe_registry"
    if SECTION_SHEET_PATTERN.search(sheet_name):
        return "section_reference"
    return None


def _normalize_pavement_status(raw_value: Any) -> str | None:
    text = _clean_text(raw_value)
    if not text:
        return None

    if "육상" in text or "수중" in text:
        return None

    upper = text.upper()
    if "ASP" in upper and "CON" in upper:
        return "콘크리트"
    if "ASP" in upper or "아스" in text:
        return "아스팔트"
    if "CON" in upper or "콘크" in text:
        return "콘크리트"
    return "비포장"


def _extract_work_environment(*values: Any) -> str | None:
    for value in values:
        text = _clean_text(value)
        if not text:
            continue
        if "수중" in text:
            return "수중"
        if "육상" in text:
            return "육상"
    return None


def detect_earthwork_sheets(workbook_path: str | Path, config: dict[str, Any] | None = None) -> list[dict[str, str]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    detected: list[dict[str, str]] = []

    for worksheet in _visible_worksheets(workbook):
        sheet_type = _sheet_type(worksheet.title)
        if sheet_type is None:
            continue
        detected.append(
            {
                "sheet_name": worksheet.title,
                "sheet_type": sheet_type,
            }
        )

    return detected


def _major_minor_to_chainage(station_major: float | None, station_minor: float | None, station_interval_m: float) -> float | None:
    if station_major is None and station_minor is None:
        return None

    major = station_major or 0.0
    minor = station_minor or 0.0
    return round(major * station_interval_m + minor, 6)


def _chainage_from_station(station_major: float | None, station_minor: float | None, station_interval_m: float) -> float | None:
    return _major_minor_to_chainage(station_major, station_minor, station_interval_m)


def _station_text(prefix: str, station_major: float | None, station_minor: float | None) -> str:
    prefix_text = prefix or "NO."
    if station_major is None and station_minor is None:
        return prefix_text

    major_text = "" if station_major is None else f"{station_major:g}"
    minor_text = "" if station_minor is None else f"{station_minor:g}"
    return f"{prefix_text}{major_text}+{minor_text}"


def _normalize_pipe_line_name(raw_name: str) -> str:
    text = _clean_text(raw_name)
    if not text:
        return ""
    if text.upper().endswith("-LINE"):
        return text
    return f"{text}-LINE"


def _infer_branch_parent(line_name: str, main_line_name: str) -> str | None:
    if line_name == main_line_name:
        return None

    stem = line_name.removesuffix("-LINE")
    if "-" not in stem:
        return main_line_name

    parent_stem = stem.rsplit("-", 1)[0]
    candidate = f"{parent_stem}-LINE"
    return candidate if candidate != line_name else main_line_name


def _normalize_earthwork_line_name(raw_name: str, main_line_name: str) -> tuple[str, str | None]:
    text = _clean_text(raw_name)
    if not text:
        return "", None

    if text.upper().endswith("-LINE"):
        return text, _infer_branch_parent(text, main_line_name)

    matched = re.match(r"^(?P<root>[A-Z]+)-(?P<branch>.+)$", text)
    if matched:
        root_name = f"{matched.group('root')}-LINE"
        normalized = f"{matched.group('branch')}-LINE"
        return normalized, root_name if normalized != root_name else _infer_branch_parent(normalized, main_line_name)

    normalized = _normalize_pipe_line_name(text)
    return normalized, _infer_branch_parent(normalized, main_line_name)


def _extract_mm_hint(raw_text: str) -> float | None:
    text = _clean_text(raw_text)
    matched_pair = DIAMETER_PAIR_PATTERN.search(text)
    if matched_pair:
        return float(matched_pair.group("inner"))

    matched = DIAMETER_HINT_PATTERN.search(text)
    if not matched:
        return None
    return float(matched.group("mm"))


def _parse_pipe_registry_workbook(pipe_workbook_path: str | Path, config: dict[str, Any], logger) -> pd.DataFrame:
    prep_config = _load_prep_config(config)
    columns = prep_config["pipe_registry_columns"]
    station_interval_m = float(prep_config["station_interval_m"])

    workbook = load_workbook(pipe_workbook_path, read_only=True, data_only=True)
    worksheet = next((ws for ws in _visible_worksheets(workbook) if PIPE_SHEET_PATTERN.search(ws.title)), None)
    if worksheet is None:
        raise EarthworkPrepError("관로공 샘플 파일에서 관수량 시트를 찾지 못했습니다.")

    records: list[dict[str, Any]] = []
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=int(prep_config["pipe_registry_start_row"]), values_only=True),
        start=int(prep_config["pipe_registry_start_row"]),
    ):
        raw_name = _clean_text(_cell_value(row, int(columns["line_name"])))
        if not raw_name or raw_name in {"소 계", "합 계"}:
            continue

        line_name = _normalize_pipe_line_name(raw_name)
        start_major = coerce_to_float(_cell_value(row, int(columns["start_major"])))
        start_minor = coerce_to_float(_cell_value(row, int(columns["start_minor"])))
        end_major = coerce_to_float(_cell_value(row, int(columns["end_major"])))
        end_minor = coerce_to_float(_cell_value(row, int(columns["end_minor"])))
        line_length_m = coerce_to_float(_cell_value(row, int(columns["line_length"])))
        if line_length_m is None or line_length_m <= 0:
            continue

        pipe_type = _clean_text(_cell_value(row, int(columns["pipe_type"])))
        diameter_text = _clean_text(_cell_value(row, int(columns["diameter_text"])))

        records.append(
            {
                "라인명": line_name,
                "원본라인명": raw_name,
                "시점측점": _station_text("NO.", start_major, start_minor),
                "종점측점": _station_text("NO.", end_major, end_minor),
                "시점거리(m)": _major_minor_to_chainage(start_major, start_minor, station_interval_m),
                "종점거리(m)": _major_minor_to_chainage(end_major, end_minor, station_interval_m),
                "연장(m)": round(line_length_m, 3),
                "관종": pipe_type or None,
                "관경표기": diameter_text or None,
                "관경추정(mm)": _extract_mm_hint(diameter_text),
                "원본행": row_number,
            }
        )

    if not records:
        raise EarthworkPrepError("관로공 파일에서 유효한 라인 정보를 찾지 못했습니다.")

    line_registry_table = pd.DataFrame(records)
    main_line_index = line_registry_table["연장(m)"].astype(float).idxmax()
    main_line_name = str(line_registry_table.loc[main_line_index, "라인명"])
    line_registry_table["메인관로명"] = main_line_name
    line_registry_table["메인관로여부"] = line_registry_table["라인명"].eq(main_line_name)
    line_registry_table["라인유형"] = line_registry_table["메인관로여부"].map({True: "메인", False: "지선"})
    line_registry_table["상위라인명"] = line_registry_table["라인명"].apply(
        lambda value: None if value == main_line_name else _infer_branch_parent(str(value), main_line_name)
    )
    line_registry_table["라인계층"] = line_registry_table["라인명"].astype(str).str.removesuffix("-LINE").str.count("-")
    line_registry_table["적용관경(mm)"] = line_registry_table["관경추정(mm)"]
    line_registry_table.sort_values(["라인명", "연장(m)"], inplace=True)
    line_registry_table.reset_index(drop=True, inplace=True)
    logger.info("관로공 파일에서 %s개 라인 정보를 읽었습니다. 메인관로=%s", len(line_registry_table), main_line_name)
    return line_registry_table


def _parse_main_profile_workbook(
    workbook_path: str | Path,
    line_registry_table: pd.DataFrame,
    config: dict[str, Any],
    logger,
) -> pd.DataFrame:
    prep_config = _load_prep_config(config)
    station_interval_m = float(prep_config["station_interval_m"])
    columns = prep_config["profile_columns"]
    profile_start_row = int(prep_config["profile_start_row"])
    visible_only = bool(prep_config.get("visible_only", True))

    main_rows = line_registry_table[line_registry_table["메인관로여부"] == True]
    if main_rows.empty:
        return pd.DataFrame()

    main_line_name = str(main_rows.iloc[0]["라인명"])
    registry_lookup = line_registry_table.set_index("라인명").to_dict("index")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheets = _visible_worksheets(workbook) if visible_only else workbook.worksheets

    records: list[dict[str, Any]] = []
    for worksheet in worksheets:
        matched = PROFILE_SHEET_PATTERN.search(worksheet.title)
        if not matched:
            continue

        profile_line_name = _normalize_pipe_line_name(matched.group("line_name"))
        if profile_line_name != main_line_name:
            continue

        registry_row = registry_lookup.get(profile_line_name)
        previous_chainage_m: float | None = None

        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=profile_start_row, values_only=True),
            start=profile_start_row,
        ):
            station_prefix = _clean_text(_cell_value(row, int(columns["station_prefix"])))
            station_major = coerce_to_float(_cell_value(row, int(columns["station_major"])))
            station_minor = coerce_to_float(_cell_value(row, int(columns["station_minor"])))
            station_note = _clean_text(_cell_value(row, int(columns["station_note"])))
            segment_length_m = coerce_to_float(_cell_value(row, int(columns["segment_length"])))
            road_width_class = _clean_text(_cell_value(row, int(columns["road_width_class"])))
            pavement_hint = _clean_text(_cell_value(row, int(columns["pavement_hint"])))

            if station_prefix and not station_prefix.upper().startswith("NO"):
                continue
            if station_major is None and station_minor is None and segment_length_m is None:
                continue
            if segment_length_m is None or segment_length_m <= 0:
                continue

            chainage_to_m = _chainage_from_station(station_major, station_minor, station_interval_m)
            if chainage_to_m is None and previous_chainage_m is not None:
                chainage_to_m = previous_chainage_m + segment_length_m
            if chainage_to_m is None:
                continue

            chainage_from_m = round(chainage_to_m - segment_length_m, 6)
            previous_chainage_m = chainage_to_m

            excavation_qty_m3 = coerce_to_float(_cell_value(row, int(columns["excavation_qty"])))
            backfill_qty_m3 = coerce_to_float(_cell_value(row, int(columns["backfill_qty"])))
            sand_qty_m3 = coerce_to_float(_cell_value(row, int(columns["sand_qty"])))
            asp_qty = coerce_to_float(_cell_value(row, int(columns["asp_qty"])))
            conc_qty = coerce_to_float(_cell_value(row, int(columns["conc_qty"])))

            pavement_status = None
            if asp_qty and conc_qty:
                pavement_status = "콘크리트"
            elif asp_qty:
                pavement_status = "아스팔트"
            elif conc_qty:
                pavement_status = "콘크리트"
            else:
                pavement_status = _normalize_pavement_status(pavement_hint)
            work_environment = _extract_work_environment(pavement_hint, station_note)

            extra_note_parts = [part for part in [station_note] if part]
            if pavement_hint and work_environment is None and pavement_status == "비포장":
                extra_note_parts.append(pavement_hint)
            elif pavement_hint and work_environment is None and pavement_status is None:
                extra_note_parts.append(pavement_hint)

            review_reasons: list[str] = []
            if registry_row is not None and chainage_to_m > float(registry_row["연장(m)"]) + 0.001:
                review_reasons.append("토적산출표 종점이 관로공 연장을 초과합니다.")

            records.append(
                {
                    "라인명": profile_line_name,
                    "상위라인명": None,
                    "라인유형": "메인",
                    "세그먼트출처": "토적산출표",
                    "원본라인명": matched.group("line_name"),
                    "원본시트": worksheet.title,
                    "원본행": row_number,
                    "관로구분": None if registry_row is None else registry_row["관경표기"],
                    "포장상태": pavement_status,
                    "시공환경": work_environment,
                    "도로폭구분": road_width_class or None,
                    "시공유형": "개착",
                    "임시점용적용": "예",
                    "영구점용적용": "예",
                    "시점측점": _station_text(station_prefix, station_major, station_minor),
                    "종점측점": _station_text(station_prefix, station_major, station_minor) + (f" {station_note}" if station_note else ""),
                    "시점거리(m)": chainage_from_m,
                    "종점거리(m)": chainage_to_m,
                    "구간연장(m)": round(segment_length_m, 3),
                    "절단연장(m)": None,
                    "차로수": None,
                    "비고": " / ".join(extra_note_parts) if extra_note_parts else None,
                    "대표단면후보": None,
                    "관로공연장(m)": None if registry_row is None else float(registry_row["연장(m)"]),
                    "토공누적연장(m)": None,
                    "원본관경표기": None if registry_row is None else registry_row["관경표기"],
                    "적용관경(mm)": None if registry_row is None else registry_row["적용관경(mm)"],
                    "터파기수량(㎥)": excavation_qty_m3,
                    "되메우기수량(㎥)": backfill_qty_m3,
                    "모래부설수량(㎥)": sand_qty_m3,
                    "검토필요": "예" if review_reasons else "아니오",
                    "검토사유": " / ".join(review_reasons) if review_reasons else None,
                }
            )

    profile_table = pd.DataFrame(records)
    logger.info("토적산출표에서 메인관로 구간 %s건을 정리했습니다.", len(profile_table))
    return profile_table


def _parse_earthwork_list_workbook(workbook_path: str | Path, line_registry_table: pd.DataFrame, config: dict[str, Any], logger) -> pd.DataFrame:
    prep_config = _load_prep_config(config)
    columns = prep_config["earthwork_list_columns"]
    station_interval_m = float(prep_config["station_interval_m"])
    visible_only = bool(prep_config.get("visible_only", True))

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheets = _visible_worksheets(workbook) if visible_only else workbook.worksheets
    registry_lookup = line_registry_table.set_index("라인명").to_dict("index")
    main_line_name = str(line_registry_table.loc[line_registry_table["메인관로여부"] == True, "라인명"].iloc[0])

    records: list[dict[str, Any]] = []
    for worksheet in worksheets:
        matched = LIST_SHEET_PATTERN.search(worksheet.title)
        if not matched:
            continue

        diameter_group = matched.group("diameter_group")
        current_line_raw = ""
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=int(prep_config["earthwork_list_start_row"]), values_only=True),
            start=int(prep_config["earthwork_list_start_row"]),
        ):
            raw_line = _clean_text(_cell_value(row, int(columns["line_name"])))
            if raw_line in {"계", "합계"}:
                break
            if raw_line:
                current_line_raw = raw_line

            pavement_status_raw = _clean_text(_cell_value(row, int(columns["pavement_status"])))
            road_width_class = _clean_text(_cell_value(row, int(columns["road_width_class"])))
            segment_length_m = coerce_to_float(_cell_value(row, int(columns["segment_length"])))
            cut_length_m = coerce_to_float(_cell_value(row, int(columns["cut_length"])))
            note_value = _clean_text(_cell_value(row, int(columns["note"])))
            lane_count = coerce_to_float(_cell_value(row, int(columns["lane_count"])))

            start_prefix = _clean_text(_cell_value(row, int(columns["start_prefix"])))
            start_major = coerce_to_float(_cell_value(row, int(columns["start_major"])))
            start_minor = coerce_to_float(_cell_value(row, int(columns["start_minor"])))
            end_prefix = _clean_text(_cell_value(row, int(columns["end_prefix"])))
            end_major = coerce_to_float(_cell_value(row, int(columns["end_major"])))
            end_minor = coerce_to_float(_cell_value(row, int(columns["end_minor"])))

            if not current_line_raw and segment_length_m is None:
                continue
            if segment_length_m is None or segment_length_m <= 0:
                continue

            normalized_line_name, parent_line_name = _normalize_earthwork_line_name(current_line_raw, main_line_name)
            start_chainage_m = _major_minor_to_chainage(start_major, start_minor, station_interval_m)
            end_chainage_m = _major_minor_to_chainage(end_major, end_minor, station_interval_m)

            registry_row = registry_lookup.get(normalized_line_name)
            review_reasons: list[str] = []
            is_pipe_hanging = pavement_status_raw == "관매달기" or "관매달기" in note_value
            pavement_status = None if is_pipe_hanging else _normalize_pavement_status(pavement_status_raw)
            work_environment = _extract_work_environment(pavement_status_raw, note_value)
            if registry_row is None:
                review_reasons.append("관로공 파일에서 대응 라인명을 찾지 못했습니다.")
            else:
                line_total_length = float(registry_row["연장(m)"])
                if end_chainage_m is not None and end_chainage_m > line_total_length + 0.001:
                    review_reasons.append("토공 구간의 종점이 관로공 연장을 초과합니다.")

            if not road_width_class and not is_pipe_hanging:
                review_reasons.append("규격 값이 비어 있습니다.")
            if not pavement_status and not is_pipe_hanging:
                review_reasons.append("포장상태 값이 비어 있습니다.")

            section_candidates = []
            if not is_pipe_hanging:
                section_candidates = _find_section_candidates(
                    workbook=workbook,
                    diameter_group=diameter_group,
                    pavement_status=pavement_status,
                    visible_only=visible_only,
                )

            records.append(
                {
                    "라인명": normalized_line_name,
                    "상위라인명": parent_line_name,
                    "라인유형": "지선",
                    "세그먼트출처": "관로토공LIST",
                    "원본라인명": current_line_raw,
                    "원본시트": worksheet.title,
                    "원본행": row_number,
                    "관로구분": diameter_group,
                    "포장상태": pavement_status or None,
                    "시공환경": work_environment,
                    "도로폭구분": road_width_class or None,
                    "시공유형": "관매달기" if is_pipe_hanging else "개착",
                    "임시점용적용": "아니오" if is_pipe_hanging else "예",
                    "영구점용적용": "예",
                    "시점측점": _station_text(start_prefix, start_major, start_minor),
                    "종점측점": _station_text(end_prefix, end_major, end_minor),
                    "시점거리(m)": start_chainage_m,
                    "종점거리(m)": end_chainage_m,
                    "구간연장(m)": round(segment_length_m, 3),
                    "절단연장(m)": round(cut_length_m, 3) if cut_length_m is not None else None,
                    "차로수": lane_count,
                    "비고": " / ".join(
                        part
                        for part in [
                            pavement_status_raw
                            if pavement_status_raw and pavement_status_raw != pavement_status and _extract_work_environment(pavement_status_raw) is None
                            else None,
                            note_value or None,
                        ]
                        if part
                    )
                    or None,
                    "대표단면후보": " | ".join(section_candidates) if section_candidates else None,
                    "관로공연장(m)": None if registry_row is None else float(registry_row["연장(m)"]),
                    "원본관경표기": None if registry_row is None else registry_row["관경표기"],
                    "적용관경(mm)": None if registry_row is None else registry_row["적용관경(mm)"],
                    "터파기수량(㎥)": None,
                    "되메우기수량(㎥)": None,
                    "모래부설수량(㎥)": None,
                    "관경표기": None if registry_row is None else registry_row["관경표기"],
                    "검토필요": "예" if review_reasons else "아니오",
                    "검토사유": " / ".join(review_reasons) if review_reasons else None,
                }
            )

    if not records:
        raise EarthworkPrepError("토공 파일의 visible 관로토공LIST 시트에서 유효한 구간을 찾지 못했습니다.")

    segment_table = pd.DataFrame(records)
    segment_table.sort_values(["라인명", "시점거리(m)", "종점거리(m)", "원본행"], inplace=True)
    segment_table.reset_index(drop=True, inplace=True)
    logger.info("토공 파일에서 %s개 관로 토공 구간을 읽었습니다.", len(segment_table))
    return segment_table


def _find_section_candidates(workbook, diameter_group: str, pavement_status: str, visible_only: bool) -> list[str]:
    worksheets = _visible_worksheets(workbook) if visible_only else workbook.worksheets
    normalized_pavement = _normalize_pavement_status(pavement_status)
    candidates: list[str] = []
    for worksheet in worksheets:
        title = worksheet.title
        if not SECTION_SHEET_PATTERN.search(title):
            continue
        if diameter_group and diameter_group not in title:
            continue

        title_upper = title.upper()
        if normalized_pavement == "아스팔트" and "ASP" not in title_upper:
            continue
        if normalized_pavement == "콘크리트" and "CON'C" not in title_upper:
            continue
        if normalized_pavement == "비포장" and "비포장" not in title and "토사" not in title:
            continue

        candidates.append(title)

    return candidates


def _parse_section_reference_table(workbook_path: str | Path, config: dict[str, Any], logger) -> pd.DataFrame:
    prep_config = _load_prep_config(config)
    visible_only = bool(prep_config.get("visible_only", True))
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheets = _visible_worksheets(workbook) if visible_only else workbook.worksheets

    records: list[dict[str, Any]] = []
    for worksheet in worksheets:
        if not SECTION_SHEET_PATTERN.search(worksheet.title):
            continue

        title = worksheet.title
        matched_parens = PARENS_PATTERN.findall(title)
        matched_bh = BH_PATTERN.search(title)
        records.append(
            {
                "시트명": title,
                "시트유형": "구조물" if "변실" in title or "가압장" in title or "유량계" in title or "공기변" in title else "관로",
                "관로구분": matched_parens[0] if matched_parens else None,
                "포장상태": _infer_pavement_from_title(title),
                "대표값": None if matched_bh is None else float(matched_bh.group("value")),
                "비고": "visible 시트만 참조",
            }
        )

    section_reference_table = pd.DataFrame(records)
    logger.info("대표단면/구조물 참고 시트 %s개를 정리했습니다.", len(section_reference_table))
    return section_reference_table


def _infer_pavement_from_title(title: str) -> str | None:
    title_upper = title.upper()
    if "ASP" in title_upper and "CON'C" in title_upper:
        return "콘크리트"
    if "ASP" in title_upper:
        return "아스팔트"
    if "CON'C" in title_upper:
        return "콘크리트"
    if "비포장" in title:
        return "비포장"
    if "토사" in title:
        return "비포장"
    return None


def _parse_structure_table(workbook_path: str | Path, config: dict[str, Any], logger) -> pd.DataFrame:
    prep_config = _load_prep_config(config)
    visible_only = bool(prep_config.get("visible_only", True))
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheets = _visible_worksheets(workbook) if visible_only else workbook.worksheets
    worksheet = next((ws for ws in worksheets if STRUCTURE_LIST_PATTERN.search(ws.title)), None)

    if worksheet is None:
        logger.info("visible 변실LIST 시트가 없어 구조물 표는 비어 있습니다.")
        return pd.DataFrame(
            columns=["구조물명", "참조라인", "위치", "도로폭구분", "포장상태", "관경추정(mm)", "수량", "원본행"]
        )

    header_condition_row = list(next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True)))
    header_diameter_row = list(next(worksheet.iter_rows(min_row=3, max_row=3, values_only=True)))

    column_specs: list[dict[str, Any]] = []
    for column_index in range(4, len(header_condition_row) + 1):
        condition_value = _clean_text(header_condition_row[column_index - 1])
        diameter_value = coerce_to_float(header_diameter_row[column_index - 1])
        if not condition_value and diameter_value is None:
            continue
        column_specs.append(
            {
                "column_index": column_index,
                "condition": condition_value or None,
                "diameter_mm": diameter_value,
            }
        )

    records: list[dict[str, Any]] = []
    current_reference_line = ""
    for row_number, row in enumerate(worksheet.iter_rows(min_row=4, values_only=True), start=4):
        first_value = _clean_text(_cell_value(row, 1))
        second_value = _clean_text(_cell_value(row, 2))
        width_class = _clean_text(_cell_value(row, 3))

        if first_value and first_value != "계":
            current_reference_line = first_value

        structure_name = second_value if second_value and "NO." not in second_value else first_value
        location_text = second_value if second_value and "NO." in second_value else None

        for column_spec in column_specs:
            quantity = coerce_to_float(_cell_value(row, int(column_spec["column_index"])))
            if quantity is None or quantity <= 0:
                continue

            records.append(
                {
                    "구조물명": structure_name or None,
                    "참조라인": current_reference_line or None,
                    "위치": location_text,
                    "도로폭구분": width_class or None,
                    "포장상태": column_spec["condition"],
                    "관경추정(mm)": column_spec["diameter_mm"],
                    "수량": quantity,
                    "원본행": row_number,
                }
            )

    structure_table = pd.DataFrame(records)
    logger.info("구조물 LIST에서 %s개 유효 행을 정리했습니다.", len(structure_table))
    return structure_table


def _find_diameter_column(columns: list[str]) -> str | None:
    aliases = {"관경", "diameter", "pipe_diameter", "dia_mm", "DIA_MM"}
    for column_name in columns:
        if column_name in aliases:
            return column_name
    return None


def _line_endpoint_keys(geometry, precision: int = 3) -> tuple[tuple[float, float] | None, tuple[float, float] | None]:
    if geometry is None or geometry.is_empty:
        return None, None

    if geometry.geom_type == "LineString":
        target = geometry
    elif geometry.geom_type == "MultiLineString":
        target = max(geometry.geoms, key=lambda item: item.length, default=None)
    else:
        return None, None

    if target is None:
        return None, None

    coordinates = list(target.coords)
    if not coordinates:
        return None, None

    start = coordinates[0]
    end = coordinates[-1]
    return (
        (round(float(start[0]), precision), round(float(start[1]), precision)),
        (round(float(end[0]), precision), round(float(end[1]), precision)),
    )


def _node_key_text(node_key: tuple[float, float] | None) -> str | None:
    if node_key is None:
        return None
    return f"{node_key[0]:.3f},{node_key[1]:.3f}"


def _summarize_pipe_features(pipe_shapefile_path: str | Path, line_registry_table: pd.DataFrame, logger) -> pd.DataFrame:
    gdf = gpd.read_file(pipe_shapefile_path)
    if gdf.empty:
        raise EarthworkPrepError("관로 SHP가 비어 있습니다.")

    diameter_column = _find_diameter_column([str(column) for column in gdf.columns])
    records: list[dict[str, Any]] = []
    for feature_index, row in gdf.iterrows():
        geometry = row.geometry
        start_node_key, end_node_key = _line_endpoint_keys(geometry)
        records.append(
            {
                "관로FID": int(feature_index),
                "도형유형": geometry.geom_type,
                "관경속성": None if diameter_column is None else row[diameter_column],
                "연장(m)": round(float(geometry.length), 3),
                "시점노드": _node_key_text(start_node_key),
                "종점노드": _node_key_text(end_node_key),
                "_start_node_key": start_node_key,
                "_end_node_key": end_node_key,
            }
        )

    pipe_feature_table = pd.DataFrame(records)
    node_to_feature_ids: dict[tuple[float, float], set[int]] = {}
    for _, row in pipe_feature_table.iterrows():
        for key_name in ["_start_node_key", "_end_node_key"]:
            node_key = row[key_name]
            if node_key is None:
                continue
            node_to_feature_ids.setdefault(node_key, set()).add(int(row["관로FID"]))

    line_registry_table = line_registry_table.copy()
    main_rows = line_registry_table[line_registry_table["메인관로여부"] == True]
    main_feature_ids: set[int] = set()
    if not main_rows.empty:
        main_length = float(main_rows.iloc[0]["연장(m)"])
        main_feature_ids = {
            int(feature_row["관로FID"])
            for _, feature_row in pipe_feature_table.iterrows()
            if abs(float(feature_row["연장(m)"]) - main_length) <= 0.001
        }

    adjacency: dict[int, set[int]] = {}
    depth_map: dict[int, int] = {}
    for _, row in pipe_feature_table.iterrows():
        feature_id = int(row["관로FID"])
        connected: set[int] = set()
        for key_name in ["_start_node_key", "_end_node_key"]:
            node_key = row[key_name]
            if node_key is None:
                continue
            connected.update(node_to_feature_ids.get(node_key, set()))
        connected.discard(feature_id)
        adjacency[feature_id] = connected

    if main_feature_ids:
        queue: deque[tuple[int, int]] = deque((feature_id, 0) for feature_id in sorted(main_feature_ids))
        visited: set[int] = set()
        while queue:
            feature_id, depth = queue.popleft()
            if feature_id in visited:
                continue
            visited.add(feature_id)
            depth_map[feature_id] = depth
            for neighbor in sorted(adjacency.get(feature_id, set())):
                if neighbor not in visited:
                    queue.append((neighbor, depth + 1))

    shape_lengths = pipe_feature_table["연장(m)"].tolist()

    candidate_ids_list: list[str | None] = []
    candidate_count_list: list[int] = []
    match_status_list: list[str] = []
    for _, line_row in line_registry_table.iterrows():
        line_length = float(line_row["연장(m)"])
        candidate_ids = [
            str(int(pipe_feature_table.iloc[index]["관로FID"]))
            for index, feature_length in enumerate(shape_lengths)
            if abs(feature_length - line_length) <= 0.001
        ]
        candidate_ids_list.append(",".join(candidate_ids) if candidate_ids else None)
        candidate_count_list.append(len(candidate_ids))
        if not candidate_ids:
            match_status_list.append("불일치")
        elif len(candidate_ids) == 1:
            match_status_list.append("단일일치")
        else:
            match_status_list.append("다중후보")

    line_registry_table["SHP후보FID"] = candidate_ids_list
    line_registry_table["SHP후보수"] = candidate_count_list
    line_registry_table["SHP매칭상태"] = match_status_list

    shared_count_values: list[int] = []
    main_connected_values: list[str] = []
    branch_depth_values: list[int | None] = []
    connected_ids_values: list[str | None] = []
    for _, feature_row in pipe_feature_table.iterrows():
        feature_id = int(feature_row["관로FID"])
        connected_ids = sorted(adjacency.get(feature_id, set()))
        shared_count_values.append(len(connected_ids))
        main_connected_values.append("예" if bool(main_feature_ids.intersection(connected_ids)) else "아니오")
        branch_depth_values.append(depth_map.get(feature_id))
        connected_ids_values.append(",".join(str(value) for value in connected_ids) if connected_ids else None)

    pipe_feature_table["접속관로수"] = shared_count_values
    pipe_feature_table["메인접속여부"] = main_connected_values
    pipe_feature_table["분기깊이"] = branch_depth_values
    pipe_feature_table["접속관로FID"] = connected_ids_values

    line_name_candidates: list[str | None] = []
    for _, feature_row in pipe_feature_table.iterrows():
        feature_length = float(feature_row["연장(m)"])
        candidate_names = line_registry_table.loc[
            (line_registry_table["연장(m)"] - feature_length).abs() <= 0.001,
            "라인명",
        ].tolist()
        line_name_candidates.append(" | ".join(candidate_names) if candidate_names else None)

    pipe_feature_table["라인후보"] = line_name_candidates

    feature_lookup = pipe_feature_table.set_index("관로FID").to_dict("index")
    candidate_summary_values: list[str | None] = []
    for _, line_row in line_registry_table.iterrows():
        candidate_ids = _clean_text(line_row.get("SHP후보FID"))
        if not candidate_ids:
            candidate_summary_values.append(None)
            continue

        summaries: list[str] = []
        for candidate_text in candidate_ids.split(","):
            feature_id = int(candidate_text)
            feature_info = feature_lookup.get(feature_id, {})
            summaries.append(
                f"FID {feature_id}(깊이={feature_info.get('분기깊이', '미확정')}, 메인접속={feature_info.get('메인접속여부', '아니오')}, 접속수={feature_info.get('접속관로수', 0)})"
            )
        candidate_summary_values.append(" / ".join(summaries))

    line_registry_table["SHP후보요약"] = candidate_summary_values
    logger.info("관로 SHP 피처 %s개를 길이 기준으로 정리했습니다.", len(pipe_feature_table))
    pipe_feature_table.drop(columns=["_start_node_key", "_end_node_key"], inplace=True)
    return line_registry_table, pipe_feature_table


def _apply_length_priority(
    segment_table: pd.DataFrame,
    line_registry_table: pd.DataFrame,
    pipe_feature_table: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    segment_table = segment_table.copy()
    line_registry_table = line_registry_table.copy()

    length_summary = segment_table.groupby("라인명", dropna=False)["구간연장(m)"].sum().round(3).to_dict()
    shp_length_lookup: dict[str, float | None] = {}
    for _, row in line_registry_table.iterrows():
        candidate_ids = row.get("SHP후보FID")
        if candidate_ids:
            first_candidate = int(str(candidate_ids).split(",")[0])
            shp_length_lookup[row["라인명"]] = float(
                pipe_feature_table.loc[pipe_feature_table["관로FID"] == first_candidate, "연장(m)"].iloc[0]
            )
        else:
            shp_length_lookup[row["라인명"]] = None

    line_registry_table["토공누적연장(m)"] = line_registry_table["라인명"].map(length_summary)
    line_registry_table["SHP연장(m)"] = line_registry_table["라인명"].map(shp_length_lookup)

    base_length_values: list[float | None] = []
    base_length_sources: list[str] = []
    length_diff_values: list[float | None] = []
    for _, row in line_registry_table.iterrows():
        shp_length = row.get("SHP연장(m)")
        pipe_length = row.get("연장(m)")
        earthwork_length = row.get("토공누적연장(m)")

        if pd.notna(shp_length):
            base_length = float(shp_length)
            source = "관로SHP"
        elif pd.notna(pipe_length):
            base_length = float(pipe_length)
            source = "관로공"
        elif pd.notna(earthwork_length):
            base_length = float(earthwork_length)
            source = "토공"
        else:
            base_length = None
            source = "미확정"

        base_length_values.append(None if base_length is None else round(base_length, 3))
        base_length_sources.append(source)
        if base_length is None or pd.isna(earthwork_length):
            length_diff_values.append(None)
        else:
            length_diff_values.append(round(float(earthwork_length) - float(base_length), 3))

    line_registry_table["기준연장(m)"] = base_length_values
    line_registry_table["기준출처"] = base_length_sources
    line_registry_table["연장차이(m)"] = length_diff_values

    line_lookup = line_registry_table.set_index("라인명")
    for column_name in ["토공누적연장(m)", "SHP연장(m)", "기준연장(m)", "기준출처", "연장차이(m)"]:
        segment_table[column_name] = segment_table["라인명"].map(line_lookup[column_name])

    return segment_table, line_registry_table


def _build_review_table(
    segment_table: pd.DataFrame,
    line_registry_table: pd.DataFrame,
    structure_table: pd.DataFrame,
    pipe_feature_table: pd.DataFrame,
) -> pd.DataFrame:
    reviews: list[dict[str, Any]] = []

    for _, row in segment_table.iterrows():
        if row["검토필요"] != "예":
            continue
        reviews.append(
            {
                "분류": "관로토공구간",
                "대상": row["라인명"],
                "위치/시트": f"{row['원본시트']}:{row['원본행']}",
                "검토사유": row["검토사유"],
            }
        )

    for _, row in line_registry_table.iterrows():
        if row["SHP매칭상태"] == "단일일치":
            pass
        else:
            reviews.append(
                {
                    "분류": "관로라인매칭",
                    "대상": row["라인명"],
                    "위치/시트": f"관로공:{int(row['원본행'])}",
                    "검토사유": f"SHP 매칭 상태={row['SHP매칭상태']} / 후보={row['SHP후보FID'] or '없음'} / 후보요약={row.get('SHP후보요약') or '없음'} / 분기점 기준 확인 필요",
                }
            )

        if pd.notna(row.get("연장차이(m)")) and abs(float(row["연장차이(m)"])) > 0.001:
            reviews.append(
                {
                    "분류": "연장불일치",
                    "대상": row["라인명"],
                    "위치/시트": f"관로공:{int(row['원본행'])}",
                    "검토사유": f"기준연장={row['기준연장(m)']}({row['기준출처']}) / 토공누적={row['토공누적연장(m)']}",
                }
            )

    if structure_table.empty:
        reviews.append(
            {
                "분류": "구조물",
                "대상": "변실LIST",
                "위치/시트": "visible 시트 미존재 또는 수량 0",
                "검토사유": "구조물 반영 대상이 있는지 확인이 필요합니다.",
            }
        )

    for _, row in pipe_feature_table.iterrows():
        if row["라인후보"]:
            continue
        reviews.append(
            {
                "분류": "관로SHP",
                "대상": f"FID {int(row['관로FID'])}",
                "위치/시트": "관로 SHP",
                "검토사유": "관로공 파일 길이와 일치하는 라인 후보를 찾지 못했습니다.",
            }
        )

    return pd.DataFrame(reviews)


def _build_final_input_table(segment_table: pd.DataFrame, config: dict[str, Any]) -> pd.DataFrame:
    defaults = get_default_selection(config)
    final_table = segment_table.copy()

    if "포장상태" in final_table.columns:
        final_table["포장상태"] = final_table["포장상태"].apply(_normalize_pavement_status)
    if "시공환경" in final_table.columns:
        final_table["시공환경"] = final_table["시공환경"].apply(_extract_work_environment)

    final_table["토피(m)"] = float(defaults.soil_cover_m)
    final_table["터파기기울기"] = str(defaults.excavation_slope)
    final_table["산정포장상태"] = final_table["포장상태"].fillna("비포장")

    def resolve_temporary_width(row) -> float | None:
        if row.get("임시점용적용") != "예":
            return 0.0

        diameter_value = row.get("적용관경(mm)")
        if pd.isna(diameter_value):
            return None

        pavement_status = row.get("산정포장상태") or "비포장"
        selection = TemporaryOccupancySelection(
            soil_cover_m=float(defaults.soil_cover_m),
            pavement_status=str(pavement_status),
            excavation_slope=str(defaults.excavation_slope),
        )
        return calculate_temporary_width(diameter_value, selection, config)

    final_table["터파기폭(m)"] = final_table.apply(resolve_temporary_width, axis=1)

    keep_columns = [
        "라인명",
        "상위라인명",
        "라인유형",
        "세그먼트출처",
        "시공유형",
        "임시점용적용",
        "영구점용적용",
        "시점거리(m)",
        "종점거리(m)",
        "구간연장(m)",
        "적용관경(mm)",
        "포장상태",
        "시공환경",
        "도로폭구분",
        "대표단면후보",
        "토피(m)",
        "터파기기울기",
        "산정포장상태",
        "터파기폭(m)",
        "기준연장(m)",
        "기준출처",
        "검토필요",
        "검토사유",
        "비고",
    ]
    available_columns = [column_name for column_name in keep_columns if column_name in final_table.columns]
    final_table = final_table.loc[:, available_columns].copy()
    final_table.sort_values(["라인유형", "라인명", "시점거리(m)", "종점거리(m)"], inplace=True)
    final_table.reset_index(drop=True, inplace=True)
    return final_table


def preprocess_earthwork_bundle(
    workbook_path: str | Path,
    pipe_workbook_path: str | Path,
    pipe_shapefile_path: str | Path,
    config: dict[str, Any],
    logger,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    line_registry_table = _parse_pipe_registry_workbook(pipe_workbook_path, config, logger)
    main_profile_table = _parse_main_profile_workbook(workbook_path, line_registry_table, config, logger)
    branch_segment_table = _parse_earthwork_list_workbook(workbook_path, line_registry_table, config, logger)
    segment_frames = [table for table in [main_profile_table, branch_segment_table] if not table.empty]
    if segment_frames:
        column_order: list[str] = []
        for table in segment_frames:
            for column_name in table.columns:
                if column_name not in column_order:
                    column_order.append(column_name)
        segment_frames = [table.reindex(columns=column_order) for table in segment_frames]
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", FutureWarning)
            segment_table = pd.concat(segment_frames, ignore_index=True, sort=False)
        segment_table.sort_values(["라인유형", "라인명", "시점거리(m)", "종점거리(m)", "원본행"], inplace=True)
        segment_table.reset_index(drop=True, inplace=True)
    else:
        raise EarthworkPrepError("토공 파일에서 메인/지선 표준 구간을 생성하지 못했습니다.")

    section_reference_table = _parse_section_reference_table(workbook_path, config, logger)
    structure_table = _parse_structure_table(workbook_path, config, logger)
    line_registry_table, pipe_feature_table = _summarize_pipe_features(pipe_shapefile_path, line_registry_table, logger)
    segment_table, line_registry_table = _apply_length_priority(segment_table, line_registry_table, pipe_feature_table)
    review_table = _build_review_table(segment_table, line_registry_table, structure_table, pipe_feature_table)
    final_input_table = _build_final_input_table(segment_table, config)

    logger.info(
        "토공 전처리를 완료했습니다. 라인 %s개, 표준 구간 %s개(메인 %s / 지선 %s), 구조물 %s개, 검토 %s건",
        len(line_registry_table),
        len(segment_table),
        len(main_profile_table),
        len(branch_segment_table),
        len(structure_table),
        len(review_table),
    )
    return (
        segment_table,
        final_input_table,
        line_registry_table,
        section_reference_table,
        structure_table,
        pipe_feature_table,
        review_table,
    )


def export_earthwork_outputs(
    segment_table: pd.DataFrame,
    final_input_table: pd.DataFrame,
    line_registry_table: pd.DataFrame,
    section_reference_table: pd.DataFrame,
    structure_table: pd.DataFrame,
    pipe_feature_table: pd.DataFrame,
    review_table: pd.DataFrame,
    output_dir: str | Path,
    config: dict[str, Any],
    logger,
) -> dict[str, str]:
    prep_config = _load_prep_config(config)
    output_config = prep_config["output"]
    csv_encoding = config["csv"]["output_encoding"]
    timestamp = now_timestamp()
    prefix = output_config["file_prefix"]

    output_path = ensure_directory(output_dir)
    runtime_temp_root = prepare_runtime_temp_directory()
    workbook_temp_dir = runtime_temp_root / "openpyxl" / f"earthwork_{timestamp}"

    workbook_path = output_path / f"{prefix}_{timestamp}_{output_config['workbook_suffix']}.xlsx"
    final_csv_path = output_path / f"{prefix}_{timestamp}_{output_config['final_csv_suffix']}.csv"
    segments_csv_path = output_path / f"{prefix}_{timestamp}_{output_config['segments_csv_suffix']}.csv"
    review_csv_path = output_path / f"{prefix}_{timestamp}_{output_config['review_csv_suffix']}.csv"
    line_registry_csv_path = output_path / f"{prefix}_{timestamp}_{output_config['line_registry_csv_suffix']}.csv"
    log_path = output_path / f"{prefix}_{timestamp}_{output_config['log_suffix']}.txt"

    try:
        with override_temp_directory(workbook_temp_dir):
            with _suppress_openpyxl_cleanup_permission_error():
                with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
                    final_input_table.to_excel(writer, sheet_name=output_config["final_sheet_name"], index=False)
                    segment_table.to_excel(writer, sheet_name=output_config["segments_sheet_name"], index=False)
                    line_registry_table.to_excel(writer, sheet_name=output_config["line_registry_sheet_name"], index=False)
                    section_reference_table.to_excel(writer, sheet_name=output_config["section_sheet_name"], index=False)
                    structure_table.to_excel(writer, sheet_name=output_config["structure_sheet_name"], index=False)
                    pipe_feature_table.to_excel(writer, sheet_name=output_config["pipe_feature_sheet_name"], index=False)
                    review_table.to_excel(writer, sheet_name=output_config["review_sheet_name"], index=False)

        final_input_table.to_csv(final_csv_path, index=False, encoding=csv_encoding)
        segment_table.to_csv(segments_csv_path, index=False, encoding=csv_encoding)
        review_table.to_csv(review_csv_path, index=False, encoding=csv_encoding)
        line_registry_table.to_csv(line_registry_csv_path, index=False, encoding=csv_encoding)

        log_text = "\n".join(get_log_messages(logger))
        log_path.write_text(log_text, encoding="utf-8")
    finally:
        shutil.rmtree(workbook_temp_dir, ignore_errors=True)

    return {
        "xlsx": str(workbook_path),
        "final_csv": str(final_csv_path),
        "segments_csv": str(segments_csv_path),
        "review_csv": str(review_csv_path),
        "line_registry_csv": str(line_registry_csv_path),
        "log_txt": str(log_path),
    }


def run_earthwork_preprocessor(inputs: EarthworkPrepInputs, config: dict[str, Any], logger) -> EarthworkPrepResult:
    workbook_path = Path(inputs.workbook_path)
    pipe_workbook_path = Path(inputs.pipe_workbook_path)
    pipe_shapefile_path = Path(inputs.pipe_shapefile_path)

    if not workbook_path.exists():
        raise EarthworkPrepError(f"토공 파일을 찾을 수 없습니다: {workbook_path}")
    if not pipe_workbook_path.exists():
        raise EarthworkPrepError(f"관로공 파일을 찾을 수 없습니다: {pipe_workbook_path}")
    if not pipe_shapefile_path.exists():
        raise EarthworkPrepError(f"관로 SHP를 찾을 수 없습니다: {pipe_shapefile_path}")

    if workbook_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise EarthworkPrepError("토공 전처리기는 xlsx 또는 xlsm 파일만 지원합니다.")
    if pipe_workbook_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise EarthworkPrepError("관로공 파일은 xlsx 또는 xlsm 형식이어야 합니다.")
    if pipe_shapefile_path.suffix.lower() != ".shp":
        raise EarthworkPrepError("관로 SHP는 .shp 파일을 선택해야 합니다.")

    (
        segment_table,
        final_input_table,
        line_registry_table,
        section_reference_table,
        structure_table,
        pipe_feature_table,
        review_table,
    ) = preprocess_earthwork_bundle(
        workbook_path=workbook_path,
        pipe_workbook_path=pipe_workbook_path,
        pipe_shapefile_path=pipe_shapefile_path,
        config=config,
        logger=logger,
    )

    output_files = export_earthwork_outputs(
        segment_table=segment_table,
        final_input_table=final_input_table,
        line_registry_table=line_registry_table,
        section_reference_table=section_reference_table,
        structure_table=structure_table,
        pipe_feature_table=pipe_feature_table,
        review_table=review_table,
        output_dir=inputs.output_dir,
        config=config,
        logger=logger,
    )

    return EarthworkPrepResult(
        segment_table=segment_table,
        final_input_table=final_input_table,
        line_registry_table=line_registry_table,
        section_reference_table=section_reference_table,
        structure_table=structure_table,
        pipe_feature_table=pipe_feature_table,
        review_table=review_table,
        output_files=output_files,
    )
