from __future__ import annotations

import shutil
from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter

from utils.helpers import ensure_directory, now_timestamp, override_temp_directory, prepare_runtime_temp_directory
from utils.logger import get_log_messages


class ExportError(Exception):
    pass


def _autosize_worksheet(worksheet) -> None:
    for column_cells in worksheet.columns:
        values = [str(cell.value) for cell in column_cells if cell.value is not None]
        if not values:
            continue
        max_length = min(max(len(value) for value in values) + 2, 40)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = max_length
    worksheet.freeze_panes = "A2"


def export_results(
    output_dir: str | Path,
    source_df: pd.DataFrame,
    data_df: pd.DataFrame,
    error_df: pd.DataFrame,
    config: dict,
    logger,
    base_name: str | None = None,
) -> dict[str, str]:
    target_dir = ensure_directory(output_dir)
    base_name = base_name or f"{config['output']['file_prefix']}_{now_timestamp()}"
    runtime_temp_root = prepare_runtime_temp_directory()
    workbook_temp_dir = runtime_temp_root / "openpyxl" / base_name

    source_csv_path = target_dir / f"{base_name}_조서원본.csv"
    data_csv_path = target_dir / f"{base_name}_DATA.csv"
    error_csv_path = target_dir / f"{base_name}_오류목록.csv"
    workbook_path = target_dir / f"{base_name}.xlsx"
    log_path = target_dir / f"{base_name}_처리로그.txt"

    try:
        source_df.to_csv(source_csv_path, index=False, encoding=config["csv"]["output_encoding"])
        data_df.to_csv(data_csv_path, index=False, encoding=config["csv"]["output_encoding"])
        error_df.to_csv(error_csv_path, index=False, encoding=config["csv"]["output_encoding"])

        sheet_names = config["output"]["workbook_sheet_names"]
        with override_temp_directory(workbook_temp_dir):
            with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
                source_df.to_excel(writer, index=False, sheet_name=sheet_names["source"])
                data_df.to_excel(writer, index=False, sheet_name=sheet_names["data"])
                error_df.to_excel(writer, index=False, sheet_name=sheet_names["error"])
                for sheet_name in writer.book.sheetnames:
                    _autosize_worksheet(writer.book[sheet_name])

        log_path.write_text("\n".join(get_log_messages(logger)), encoding="utf-8")
    except Exception as error:
        raise ExportError(f"출력 파일 저장에 실패했습니다: {error}") from error
    finally:
        shutil.rmtree(workbook_temp_dir, ignore_errors=True)

    return {
        "source_csv": str(source_csv_path),
        "data_csv": str(data_csv_path),
        "error_csv": str(error_csv_path),
        "xlsx": str(workbook_path),
        "log_txt": str(log_path),
    }
