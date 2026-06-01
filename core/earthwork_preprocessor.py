from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from contextlib import contextmanager
import re
import shutil
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet._writer import ALL_TEMP_FILES, WorksheetWriter
import pandas as pd

from utils.helpers import (
    coerce_to_float,
    ensure_directory,
    now_timestamp,
    override_temp_directory,
    prepare_runtime_temp_directory,
)
from utils.logger import get_log_messages


class EarthworkPreprocessorError(ValueError):
    pass


ROAD_WIDTH_RANGE_PATTERN = re.compile(
    r"(?P<lower>\d+(?:\.\d+)?)\s*M?\s*[~\-]\s*(?P<upper>\d+(?:\.\d+)?)\s*M?",
    re.IGNORECASE,
)


@dataclass
class EarthworkPrepInputs:
    workbook_path: str
    output_dir: str


@dataclass
class EarthworkPrepResult:
    segment_table: pd.DataFrame
    summary_table: pd.DataFrame
    review_table: pd.DataFrame
    output_files: dict[str, str]


@contextmanager
def _suppress_openpyxl_cleanup_permission_error():
    original_cleanup = WorksheetWriter.cleanup

    def safe_cleanup(self) -> None:
        try:
            original_cleanup(self)
        except PermissionError:
            if getattr(self, "out", None) in ALL_TEMP_FILES:
                try:
                    ALL_TEMP_FILES.remove(self.out)
                except ValueError:
                    pass

    WorksheetWriter.cleanup = safe_cleanup
    try:
        yield
    finally:
        WorksheetWriter.cleanup = original_cleanup


def _load_prep_config(config: dict[str, Any]) -> dict[str, Any]:
    prep_config = config.get("earthwork_preprocessor")
    if not isinstance(prep_config, dict):
        raise EarthworkPreprocessorError("earthwork_preprocessor 설정이 없습니다.")
    return prep_config


def _compile_sheet_patterns(patterns: list[str]) -> list[re.Pattern[str]]:
    compiled_patterns: list[re.Pattern[str]] = []
    for pattern in patterns:
        compiled_patterns.append(re.compile(pattern))
    return compiled_patterns


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _cell_value(row: tuple[Any, ...], one_based_index: int) -> Any:
    zero_based_index = one_based_index - 1
    if zero_based_index < 0 or zero_based_index >= len(row):
        return None
    return row[zero_based_index]


def _parse_road_width_range(raw_text: str) -> tuple[float | None, float | None]:
    text = raw_text.replace(" ", "")
    if not text:
        return None, None

    matched = ROAD_WIDTH_RANGE_PATTERN.search(text)
    if matched:
        return (
            float(matched.group("lower")),
            float(matched.group("upper")),
        )

    return None, None


def _infer_shoring_required_from_workbook() -> tuple[str, str]:
    return (
        "검토필요",
        "현재 샘플 토공파일의 규격 값은 굴착깊이가 아니라 도로폭/차로폭 구분으로 보여 자동 가시설 판단을 보류했습니다.",
    )


def _infer_earthwork_method(shoring_required: str) -> str:
    if shoring_required == "예":
        return "가시설"
    if shoring_required == "아니오":
        return "자연터파기"
    return "검토필요"


def _sum_positive(values: list[float | None]) -> float:
    total = 0.0
    for value in values:
        if value is None:
            continue
        if value > 0:
            total += value
    return total


def _infer_pavement_status(
    asphalt_values: list[float | None],
    concrete_values: list[float | None],
) -> tuple[str, str | None]:
    asphalt_total = _sum_positive(asphalt_values)
    concrete_total = _sum_positive(concrete_values)

    if asphalt_total > 0 and concrete_total > 0:
        return "복합포장", "아스팔트와 콘크리트 관련 수량이 동시에 존재합니다."
    if asphalt_total > 0:
        return "아스팔트", None
    if concrete_total > 0:
        return "콘크리트", None
    return "비포장", None


def _compose_station_text(
    station_prefix: str,
    station_major: float | None,
    station_minor: float | None,
    station_note: str,
) -> str:
    prefix = station_prefix or "NO."
    if station_major is None and station_minor is None:
        return station_note

    major_text = "" if station_major is None else f"{station_major:g}"
    minor_text = "" if station_minor is None else f"{station_minor:g}"
    station_text = f"{prefix}{major_text}+{minor_text}"
    if station_note:
        station_text = f"{station_text} {station_note}"
    return station_text.strip()


def _chainage_from_station(
    station_major: float | None,
    station_minor: float | None,
    station_interval_m: float,
) -> float | None:
    if station_major is None and station_minor is None:
        return None
    major = station_major or 0.0
    minor = station_minor or 0.0
    return round(major * station_interval_m + minor, 6)


def _build_representative_section(
    site_category: str,
    pavement_status: str,
    road_width_class_raw: str,
) -> str:
    parts = [part for part in [site_category, pavement_status, road_width_class_raw] if part]
    return " / ".join(parts)


def _extract_line_name(sheet_name: str, compiled_patterns: list[re.Pattern[str]]) -> str | None:
    for pattern in compiled_patterns:
        matched = pattern.search(sheet_name)
        if not matched:
            continue
        group_dict = matched.groupdict()
        if "line_name" in group_dict:
            return _clean_text(group_dict["line_name"])
        return _clean_text(matched.group(0))
    return None


def detect_earthwork_sheets(workbook_path: str | Path, config: dict[str, Any]) -> list[dict[str, str]]:
    prep_config = _load_prep_config(config)
    compiled_patterns = _compile_sheet_patterns(prep_config["sheet_name_patterns"])
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)

    detected: list[dict[str, str]] = []
    for sheet_name in workbook.sheetnames:
        line_name = _extract_line_name(sheet_name, compiled_patterns)
        if line_name is None:
            continue
        detected.append(
            {
                "sheet_name": sheet_name,
                "line_name": line_name,
            }
        )
    return detected


def _parse_sheet_records(
    worksheet,
    sheet_name: str,
    line_name: str,
    config: dict[str, Any],
    logger,
) -> list[dict[str, Any]]:
    prep_config = _load_prep_config(config)
    columns = prep_config["columns"]
    data_start_row = int(prep_config["data_start_row"])
    station_interval_m = float(prep_config["station_interval_m"])
    records: list[dict[str, Any]] = []
    previous_chainage_m: float | None = None

    for excel_row_number, row in enumerate(
        worksheet.iter_rows(min_row=data_start_row, values_only=True),
        start=data_start_row,
    ):
        station_prefix = _clean_text(_cell_value(row, int(columns["station_prefix"])))
        station_major = coerce_to_float(_cell_value(row, int(columns["station_major"])))
        station_minor = coerce_to_float(_cell_value(row, int(columns["station_minor"])))
        station_note = _clean_text(_cell_value(row, int(columns["station_note"])))
        segment_length_m = coerce_to_float(_cell_value(row, int(columns["segment_length"])))
        road_width_class_raw = _clean_text(_cell_value(row, int(columns["equipment_class"])))
        site_category = _clean_text(_cell_value(row, int(columns["site_category"])))

        excavation_qty_m3 = coerce_to_float(_cell_value(row, int(columns["excavation_qty"])))
        backfill_qty_m3 = coerce_to_float(_cell_value(row, int(columns["backfill_qty"])))
        sand_qty_m3 = coerce_to_float(_cell_value(row, int(columns["sand_qty"])))
        asp_break_qty = coerce_to_float(_cell_value(row, int(columns["asp_break_qty"])))
        conc_break_qty = coerce_to_float(_cell_value(row, int(columns["conc_break_qty"])))
        asp_milling_qty = coerce_to_float(_cell_value(row, int(columns["asp_milling_qty"])))
        subbase_qty = coerce_to_float(_cell_value(row, int(columns["subbase_qty"])))
        asp_restore_qty = coerce_to_float(_cell_value(row, int(columns["asp_restore_qty"])))
        conc_restore_qty = coerce_to_float(_cell_value(row, int(columns["conc_restore_qty"])))
        asp_cut_qty = coerce_to_float(_cell_value(row, int(columns["asp_cut_qty"])))
        conc_cut_qty = coerce_to_float(_cell_value(row, int(columns["conc_cut_qty"])))

        has_station = station_major is not None or station_minor is not None
        has_payload = any(
            [
                segment_length_m is not None,
                bool(road_width_class_raw),
                excavation_qty_m3 is not None,
                backfill_qty_m3 is not None,
                sand_qty_m3 is not None,
                asp_break_qty is not None,
                conc_break_qty is not None,
                asp_milling_qty is not None,
                subbase_qty is not None,
                asp_restore_qty is not None,
                conc_restore_qty is not None,
                asp_cut_qty is not None,
                conc_cut_qty is not None,
            ]
        )
        if not has_station and not has_payload:
            continue

        chainage_to_m = _chainage_from_station(station_major, station_minor, station_interval_m)
        if segment_length_m is None and chainage_to_m is not None and previous_chainage_m is not None:
            segment_length_m = round(max(chainage_to_m - previous_chainage_m, 0.0), 6)

        if chainage_to_m is None and previous_chainage_m is not None and segment_length_m is not None:
            chainage_to_m = previous_chainage_m + segment_length_m

        chainage_from_m = None
        if chainage_to_m is not None and segment_length_m is not None:
            chainage_from_m = round(chainage_to_m - segment_length_m, 6)
        elif previous_chainage_m is not None:
            chainage_from_m = previous_chainage_m

        if chainage_to_m is not None:
            previous_chainage_m = chainage_to_m

        if segment_length_m is None or segment_length_m <= 0:
            continue

        road_width_min_m, road_width_max_m = _parse_road_width_range(road_width_class_raw)
        shoring_required, shoring_reason = _infer_shoring_required_from_workbook()
        pavement_status, pavement_reason = _infer_pavement_status(
            asphalt_values=[asp_break_qty, asp_milling_qty, asp_restore_qty, asp_cut_qty],
            concrete_values=[conc_break_qty, conc_restore_qty, conc_cut_qty],
        )
        representative_section = _build_representative_section(site_category, pavement_status, road_width_class_raw)

        review_reasons: list[str] = []
        if not road_width_class_raw:
            review_reasons.append("도로폭/차로폭 구분 값이 비어 있습니다.")
        if shoring_reason:
            review_reasons.append(shoring_reason)
        if pavement_reason:
            review_reasons.append(pavement_reason)
        if chainage_from_m is None or chainage_to_m is None:
            review_reasons.append("측점으로부터 구간 chainage를 확정하지 못했습니다.")

        station_text = _compose_station_text(
            station_prefix=station_prefix,
            station_major=station_major,
            station_minor=station_minor,
            station_note=station_note,
        )

        record = {
            "라인명": line_name,
            "원본시트": sheet_name,
            "원본행": excel_row_number,
            "측점표기": station_text,
            "측점주번호": station_major,
            "측점보조번호": station_minor,
            "측점메모": station_note,
            "시점거리(m)": chainage_from_m,
            "종점거리(m)": chainage_to_m,
            "구간연장(m)": round(segment_length_m, 3),
            "도로폭구분": road_width_class_raw,
            "도로폭최소(m)": road_width_min_m,
            "도로폭최대(m)": road_width_max_m,
            "가시설필요": shoring_required,
            "터파기방식": _infer_earthwork_method(shoring_required),
            "포장상태": pavement_status,
            "대표단면": representative_section,
            "부지구분": site_category or None,
            "터파기수량(㎥)": excavation_qty_m3,
            "되메우기수량(㎥)": backfill_qty_m3,
            "모래부설수량(㎥)": sand_qty_m3,
            "아스팔트깨기수량": asp_break_qty,
            "콘크리트깨기수량": conc_break_qty,
            "아스팔트절삭수량": asp_milling_qty,
            "보조기층수량": subbase_qty,
            "아스팔트복구수량": asp_restore_qty,
            "콘크리트복구수량": conc_restore_qty,
            "아스팔트절단수량": asp_cut_qty,
            "콘크리트절단수량": conc_cut_qty,
            "검토필요": "예" if review_reasons else "아니오",
            "검토사유": " / ".join(review_reasons) if review_reasons else None,
        }
        records.append(record)

    logger.info("%s 시트에서 %s개 표준 구간을 추출했습니다.", sheet_name, len(records))
    return records


def preprocess_earthwork_workbook(workbook_path: str | Path, config: dict[str, Any], logger) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    workbook_path = str(workbook_path)
    detected_sheets = detect_earthwork_sheets(workbook_path, config)
    if not detected_sheets:
        raise EarthworkPreprocessorError("토적산출표(*-LINE) 형식의 시트를 찾지 못했습니다.")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    all_records: list[dict[str, Any]] = []

    for sheet_info in detected_sheets:
        worksheet = workbook[sheet_info["sheet_name"]]
        all_records.extend(
            _parse_sheet_records(
                worksheet=worksheet,
                sheet_name=sheet_info["sheet_name"],
                line_name=sheet_info["line_name"],
                config=config,
                logger=logger,
            )
        )

    if not all_records:
        raise EarthworkPreprocessorError("표준화할 토공 구간을 찾지 못했습니다.")

    segment_table = pd.DataFrame(all_records)
    segment_table.sort_values(["라인명", "시점거리(m)", "종점거리(m)", "원본행"], inplace=True)
    segment_table.reset_index(drop=True, inplace=True)

    summary_table = (
        segment_table.groupby("라인명", dropna=False)
        .agg(
            원본시트=("원본시트", "first"),
            구간수=("라인명", "size"),
            총연장_m=("구간연장(m)", "sum"),
            가시설연장_m=("구간연장(m)", lambda series: float(segment_table.loc[series.index, "구간연장(m)"][segment_table.loc[series.index, "가시설필요"] == "예"].sum())),
            자연터파기연장_m=("구간연장(m)", lambda series: float(segment_table.loc[series.index, "구간연장(m)"][segment_table.loc[series.index, "가시설필요"] == "아니오"].sum())),
            검토필요건수=("검토필요", lambda series: int((series == "예").sum())),
        )
        .reset_index()
    )
    summary_table["총연장_m"] = summary_table["총연장_m"].round(3)
    summary_table["가시설연장_m"] = summary_table["가시설연장_m"].round(3)
    summary_table["자연터파기연장_m"] = summary_table["자연터파기연장_m"].round(3)

    review_table = segment_table[segment_table["검토필요"] == "예"].copy()
    review_table.reset_index(drop=True, inplace=True)

    logger.info(
        "토공 전처리를 완료했습니다. 라인 %s개, 구간 %s개, 검토필요 %s건",
        summary_table.shape[0],
        segment_table.shape[0],
        review_table.shape[0],
    )
    return segment_table, summary_table, review_table


def export_earthwork_outputs(
    segment_table: pd.DataFrame,
    summary_table: pd.DataFrame,
    review_table: pd.DataFrame,
    output_dir: str | Path,
    config: dict[str, Any],
    logger,
) -> dict[str, str]:
    prep_config = _load_prep_config(config)
    output_config = prep_config["output"]
    csv_encoding = config["csv"]["output_encoding"]
    timestamp = now_timestamp()
    prefix = output_config["file_prefix"]

    output_path = ensure_directory(output_dir)
    runtime_temp_root = prepare_runtime_temp_directory()
    workbook_temp_dir = runtime_temp_root / "openpyxl" / f"earthwork_{timestamp}"
    workbook_path = output_path / f"{prefix}_{timestamp}_{output_config['workbook_suffix']}.xlsx"
    segments_csv_path = output_path / f"{prefix}_{timestamp}_{output_config['segments_csv_suffix']}.csv"
    review_csv_path = output_path / f"{prefix}_{timestamp}_{output_config['review_csv_suffix']}.csv"
    log_path = output_path / f"{prefix}_{timestamp}_{output_config['log_suffix']}.txt"

    try:
        with override_temp_directory(workbook_temp_dir):
            with _suppress_openpyxl_cleanup_permission_error():
                with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
                    segment_table.to_excel(writer, sheet_name=output_config["segments_sheet_name"], index=False)
                    summary_table.to_excel(writer, sheet_name=output_config["summary_sheet_name"], index=False)
                    review_table.to_excel(writer, sheet_name=output_config["review_sheet_name"], index=False)

        segment_table.to_csv(segments_csv_path, index=False, encoding=csv_encoding)
        review_table.to_csv(review_csv_path, index=False, encoding=csv_encoding)

        log_text = "\n".join(get_log_messages(logger))
        log_path.write_text(log_text, encoding="utf-8")
    finally:
        shutil.rmtree(workbook_temp_dir, ignore_errors=True)

    return {
        "xlsx": str(workbook_path),
        "segments_csv": str(segments_csv_path),
        "review_csv": str(review_csv_path),
        "log_txt": str(log_path),
    }


def run_earthwork_preprocessor(inputs: EarthworkPrepInputs, config: dict[str, Any], logger) -> EarthworkPrepResult:
    workbook_path = Path(inputs.workbook_path)
    if not workbook_path.exists():
        raise EarthworkPreprocessorError(f"토공 엑셀 파일을 찾을 수 없습니다: {workbook_path}")

    if workbook_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise EarthworkPreprocessorError("토공 전처리기는 xlsx 또는 xlsm 파일만 지원합니다.")

    segment_table, summary_table, review_table = preprocess_earthwork_workbook(workbook_path, config, logger)
    output_files = export_earthwork_outputs(segment_table, summary_table, review_table, inputs.output_dir, config, logger)
    return EarthworkPrepResult(
        segment_table=segment_table,
        summary_table=summary_table,
        review_table=review_table,
        output_files=output_files,
    )
